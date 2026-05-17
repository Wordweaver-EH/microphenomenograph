#!/usr/bin/env python3
"""
Convert OSF XLSX analysis files to markdown format for use as few-shot examples.

Usage:
    python convert_osf_analyses.py <input_dir> <output_dir>

Example:
    python convert_osf_analyses.py osf-archive/Phase\ 1/analyses/ \
        microphenomenograph/1.0.0/examples/analyses/phase1/

The OSF XLSX files contain three sheets: 'Diachronic Analysis', 'Diachronic Structure',
and 'Synchronic Analysis'. All three are parsed. The 'Diachronic Structure' sheet contains
IDUs in experiential order with a 'Hinge' row annotating the transition criterion between
each adjacent IDU pair — this is included in the output markdown as a ## Diachronic Structure
table, enabling benchmarking against OSF reference data.
"""
import sys
import re
from pathlib import Path
import openpyxl


def clean(text):
    """Normalise smart quotes and strip whitespace."""
    if text is None:
        return ""
    return (
        str(text)
        .replace("'", "'").replace("'", "'")
        .replace(""", '"').replace(""", '"')
        .replace("�", "?")
        .strip()
    )


def parse_diachronic(ws):
    """
    Parse 'Diachronic Analysis - Participa' sheet.

    Returns list of dicts:
      { idu_name, idu_number, utterance_numbers, criteria }
    where idu_number is assigned sequentially.
    """
    rows = list(ws.iter_rows(values_only=True))
    # Row 0: header text; Row 1: column names; Rows 2+: data
    idus = []
    current_idu = None
    idu_counter = 0

    for row in rows[2:]:
        speaker, num, utterance, moment, idu_name, criteria = row[:6]
        idu_name = clean(idu_name)
        if idu_name:
            if current_idu:
                idus.append(current_idu)
            idu_counter += 1
            current_idu = {
                "idu_number": idu_counter,
                "idu_name": idu_name,
                "moment": moment,
                "utterance_numbers": [str(num)] if num is not None else [],
                "criteria": clean(criteria),
            }
        elif current_idu and num is not None:
            current_idu["utterance_numbers"].append(str(num))

    if current_idu:
        idus.append(current_idu)

    return idus


def parse_synchronic(ws):
    """
    Parse 'Synchronic Analysis' sheet.

    Returns list of dicts:
      { idu_name, utterance_numbers, criteria, isus: [{isu_name, isu_2nd_level}] }

    Handles both 5-column (no ISU 2nd Level) and 6-column formats.
    """
    rows = list(ws.iter_rows(values_only=True))
    # Row 0: "Table 1"; Row 1: column names; Rows 2+: data
    groups = []
    current_group = None

    for row in rows[2:]:
        # Handle variable column count: some files have 5 cols, some have 6
        if len(row) < 5:
            continue
        idu_name = row[0]
        num = row[1]
        utterance = row[2]
        criteria = row[3]
        isu = row[4]
        isu_2nd = row[5] if len(row) > 5 else None

        idu_name = clean(idu_name)
        isu = clean(isu)
        isu_2nd = clean(isu_2nd)
        criteria_clean = clean(criteria)

        if idu_name:
            if current_group:
                groups.append(current_group)
            current_group = {
                "idu_name": idu_name,
                "utterance_numbers": [str(num)] if num is not None else [],
                "isus": [],
            }
            if isu:
                current_group["isus"].append({
                    "isu_name": isu,
                    "isu_2nd_level": isu_2nd,
                    "criteria": criteria_clean,
                })
        elif current_group:
            if num is not None:
                current_group["utterance_numbers"].append(str(num))
            if isu:
                current_group["isus"].append({
                    "isu_name": isu,
                    "isu_2nd_level": isu_2nd,
                    "criteria": criteria_clean,
                })

    if current_group:
        groups.append(current_group)

    return groups


def parse_diachronic_structure(ws):
    """
    Parse 'Diachronic Structure' sheet.

    Sheet layout (wide format):
      Row 1 (index 0): "Table 1"
      Row 2 (index 1): IDU names in alternating columns (cols 1, 3, 5, ...)
      Row 3 (index 2): Hinges in alternating columns (cols 2, 4, 6, ...)

    Returns list of dicts: { idu_name, hinge_to_next }
    where hinge_to_next is None for the last IDU.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    idu_row = rows[1]   # IDU names
    hinge_row = rows[2]  # Hinge criteria

    # IDU names are at columns 1, 3, 5, ... (0-indexed); hinges at 2, 4, 6, ...
    idus = []
    col = 1
    while col < len(idu_row):
        name = clean(idu_row[col]) if col < len(idu_row) else ""
        if not name:
            break
        hinge = clean(hinge_row[col + 1]) if col + 1 < len(hinge_row) else ""
        idus.append({
            "idu_name": name,
            "hinge_to_next": hinge if hinge else None,
        })
        col += 2

    # Last IDU has no hinge
    if idus:
        idus[-1]["hinge_to_next"] = None

    return idus


def to_diachronic_md(header, idus, structure=None):
    lines = [f"# {header}", "", "## Diachronic Analysis", ""]
    lines.append("| IDU # | IDU Name | Utterance Numbers | Criteria |")
    lines.append("|---|---|---|---|")
    for idu in idus:
        utts = ", ".join(idu["utterance_numbers"])
        lines.append(
            f"| {idu['idu_number']} | {idu['idu_name']} "
            f"| {utts} | {idu['criteria']} |"
        )
    if structure and len(structure) >= 2:
        lines += ["", "## Diachronic Structure", ""]
        lines.append("| IDU | Hinge | IDU |")
        lines.append("|---|---|---|")
        for i in range(len(structure) - 1):
            left = structure[i]["idu_name"]
            hinge = structure[i]["hinge_to_next"] or ""
            right = structure[i + 1]["idu_name"]
            lines.append(f"| {left} | {hinge} | {right} |")
    return "\n".join(lines) + "\n"


def to_synchronic_md(header, groups):
    lines = [f"# {header}", "", "## Synchronic Analysis", ""]
    lines.append("| IDU Name | ISU Name | ISU 2nd Level | Utterance Numbers | Criteria |")
    lines.append("|---|---|---|---|---|")
    for group in groups:
        utts = ", ".join(group["utterance_numbers"])
        if group["isus"]:
            for i, isu in enumerate(group["isus"]):
                idu_cell = group["idu_name"] if i == 0 else ""
                utts_cell = utts if i == 0 else ""
                lines.append(
                    f"| {idu_cell} | {isu['isu_name']} "
                    f"| {isu['isu_2nd_level']} | {utts_cell} | {isu['criteria']} |"
                )
        else:
            lines.append(f"| {group['idu_name']} | | | {utts} | |")
    return "\n".join(lines) + "\n"


def convert_file(xlsx_path, out_dir):
    stem = xlsx_path.stem  # e.g. "p1s1"
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Diachronic sheet (may be misspelled as "Diachonic")
    dia_sheet_name = next(
        (s for s in wb.sheetnames if "diachronic" in s.lower() or "diachonic" in s.lower()), None
    )
    structure = []
    struct_sheet_name = next(
        (s for s in wb.sheetnames if ("diachronic" in s.lower() or "diachonic" in s.lower()) and "structure" in s.lower()), None
    )
    if struct_sheet_name:
        structure = parse_diachronic_structure(wb[struct_sheet_name])
    if dia_sheet_name:
        ws = wb[dia_sheet_name]
        header = clean(list(ws.iter_rows(values_only=True))[0][0])
        idus = parse_diachronic(ws)
        md = to_diachronic_md(header, idus, structure)
        (out_dir / f"{stem}-diachronic.md").write_text(md, encoding="utf-8")

    # Synchronic sheet (may be named "Synchronic Analysis" or "Synchronic Structure")
    syn_sheet_name = next(
        (s for s in wb.sheetnames if "synchronic" in s.lower() and "analysis" in s.lower()), None
    )
    if not syn_sheet_name:
        # Fallback for files that call it "Synchronic Structure"
        syn_sheet_name = next(
            (s for s in wb.sheetnames if "synchronic" in s.lower()), None
        )
    if syn_sheet_name:
        ws = wb[syn_sheet_name]
        header_row = list(ws.iter_rows(values_only=True))[0][0]
        # Use diachronic header for consistency (same participant)
        header = clean(header_row) if header_row and header_row != "Table 1" else ""
        if not header and dia_sheet_name:
            ws_dia = wb[dia_sheet_name]
            header = clean(list(ws_dia.iter_rows(values_only=True))[0][0])
        groups = parse_synchronic(wb[syn_sheet_name])
        md = to_synchronic_md(header, groups)
        (out_dir / f"{stem}-synchronic.md").write_text(md, encoding="utf-8")


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    in_dir = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted(in_dir.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No XLSX files found in {in_dir}", file=sys.stderr)
        sys.exit(1)

    for xlsx_path in xlsx_files:
        print(f"Converting {xlsx_path.name}...")
        convert_file(xlsx_path, out_dir)

    print(f"Done. {len(xlsx_files)} files converted to {out_dir}")


if __name__ == "__main__":
    main()
